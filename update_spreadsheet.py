# Author: Michael Cope
# Date: 6-20-20
# Purpose: Adds resume/job application files from a specified directory to a spreadsheet.

# Works with Ubuntu and Windows, but hasn't been tested on Mac. I assume it should though.

# This file assumes that a worksheet already exists in the specified location and has the first row of the first and
# second column occupied with "Name", and "Date" respectively. There are no real consequences if these fields aren't
# filled in however. If the workbook doesn't exist, or if the path to it isn't correct, then the program will fail with
# FileNotFoundError: [Errno 2] No such file or directory: ...

# For this app to function correctly, there are two statements you must change to reflect where your job application
# folder is, and where the spreadsheet file is at. These lines have an ALLCAPS comment above them stating this.

import openpyxl
import os
import re
from datetime import datetime


def update_spreadsheet():

    # Here is the path for the spreadsheet file that will store our job applications history
    # FILL IN ONE OF THESE LINES AND UNCOMMENT IT.
    # workbook_path = "PATH/TO/WORKBOOK.xlsx"
    # workbook_path = "C:\\Users\\YOUR_USERNAME\\Desktop_or_wherever\\SPREADSHEET.xlsx"

    wb = openpyxl.load_workbook(workbook_path)

    # You can set the specific sheet that you want to save the data to or the last active sheet.
    # Only one of the two following lines should be uncommented.
    # job_app_sheet = wb["Sheet 4"]
    job_app_sheet = wb.active

    # grabs names of the resumes/applications that are already in the spreadsheet to prevent duplication, or rewriting
    # the same data multiple times.
    old_apps = []
    for i in range(1, job_app_sheet.max_row+1):
        old_apps.append(job_app_sheet.cell(row=i, column=1).value)

    # Removes the first cell as it will be occupied by the column title.
    old_apps = old_apps[1:]

    # Finds the first row that is empty, and sets index 1 past it to avoid overwriting previous entries.
    new_index = job_app_sheet.max_row + 1

    # Here is the path for the resume/job application directory.
    # FILL IN ONE OF THESE LINES AND UNCOMMENT IT.
    # job_apps_path = "PATH/TO/JOB/APPLICATIONS/DIRECTORY"
    # job_apps_path = "C:\\Users\\YOUR_USERNAME\\Desktop_or_wherever\\JOB-APPS-FOLDER"

    # Here is where the worksheet is updated.
    for app_name in os.listdir(job_apps_path):
        # First step is to clean the file names in the resume folder
        cleaned_app_name = file_extension_remover(app_name)

        # Next, each clean name is compared against the ones already in the excel sheet.
        if cleaned_app_name not in old_apps:

            # If it isn't in the spreadsheet, the cleaned name gets added, along with the last modified date and time.
            job_app_sheet.cell(row=new_index, column=1).value = cleaned_app_name

            full_path = os.path.join(job_apps_path, app_name)

            date_time = datetime.fromtimestamp(os.path.getmtime(full_path))
            human_time = date_time.strftime("%m/%d/%Y, %H:%M:%S")

            job_app_sheet.cell(row=new_index, column=2).value = human_time
            # Moves index of spreadsheet down to avoid overwriting.
            new_index += 1

    # At the end of the cycle, the worksheet is saved to preserve changes and is closed.
    wb.save(workbook_path)
    wb.close()


def file_extension_remover(file_name):
    # This util function removes the file extension from the file name if it is present. Otherwise the name is returned.

    # More specific patterns if you want to parse files with multiple "." in the name
    # pattern = re.compile(r'\.pdf')
    # pattern = re.compile(r'\.doc')
    # pattern = re.compile(r'\.docx')
    # pattern = re.compile(r'\.odt')

    # Searches for any periods followed by one or more chars. It essentially matches any filename extension.
    pattern = re.compile(r'\..+')

    pattern_match = pattern.search(file_name)

    # Checks if a match was found.
    # If so, it returns the file name up to the beginning of the match.
    if bool(pattern_match):
        return file_name[:pattern_match.span()[0]]

    else:
        return file_name


def main():
    # Runs the updating function
    update_spreadsheet()
    # Additional functionality could be added here as well, perhaps to backup the workbook to Dropbox or Google Drive.


if __name__ == '__main__':
    main()
